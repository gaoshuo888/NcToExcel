import netCDF4
import pandas as pd
from netCDF4 import num2date
import numpy as np
from datetime import datetime, timedelta
import cftime

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 文件路径
nc_file_path = 'E:\\work\\9ProjectTyphoon\\03RainfallDataWithNc\\1985-2024\\1988\\198805.nc'
excel_file_path = 'E:\\work\\9ProjectTyphoon\\05RainfallDataWithExcel\\198805.xlsx'

# UTC+8 时间范围
start_time_utc8 = '1988-07-11 00:00:00'  # 开始时间 (UTC+8)
end_time_utc8 = '1988-07-19 12:00:00'    # 结束时间 (UTC+8)

# 经度和纬度范围
lon_min, lon_max = 110.041, 111.212  # 经度范围
lat_min, lat_max = 22.828, 24.124    # 纬度范围

#------------------------------------------------------------------------------------------------------------
cell_locations = {}
label_number = 1

for row in range(2, 7):  # 从第2行到第7行
    for col in range(ord('B'), ord('F') + 1):  # 从列B到列F
        cell = chr(col) + str(row)
        cell_locations[cell] = f'Rainfall{label_number:02d}'
        label_number += 1

print(cell_locations)
# --------------------------------------------------------------------
# 自定义第一列的标题
first_column_title = 'Time'

# 读取 Excel 文件
input_file = excel_file_path
# 保存汇总结果到新的 Excel 文件
output_file = 'E:\\work\\9ProjectTyphoon\\05RainfallDataWithExcel\\summary_output.xlsx'
#output_file = input_file

# 打开nc文件
dataset = netCDF4.Dataset(nc_file_path)

# 获取时间、经度和纬度变量
time_var = dataset.variables['time']
time = time_var[:]
longitude = dataset.variables['longitude'][:]
latitude = dataset.variables['latitude'][:]
tp = dataset.variables['tp'][:]  # (time, latitude, longitude)

# 获取时间单位
time_units = time_var.units

# 将时间变量转换为 UTC 时间
time_dates_utc = num2date(time, units=time_units, calendar='gregorian')

# 定义时间偏移量
time_offset = timedelta(hours=8)  # UTC+8

# 将 UTC+8 时间范围转换为 UTC 时间范围
start_time_utc8_dt = datetime.strptime(start_time_utc8, '%Y-%m-%d %H:%M:%S')
end_time_utc8_dt = datetime.strptime(end_time_utc8, '%Y-%m-%d %H:%M:%S')

start_time_utc_dt = start_time_utc8_dt - time_offset
end_time_utc_dt = end_time_utc8_dt - time_offset

# 将转换后的时间范围转换为与 netCDF4 时间变量兼容的时间对象
start_time_utc = cftime.date2num(start_time_utc_dt, units=time_units, calendar='gregorian')
end_time_utc = cftime.date2num(end_time_utc_dt, units=time_units, calendar='gregorian')

# 找到时间范围内的索引
time_indices = np.where((time >= start_time_utc) & (time <= end_time_utc))[0]

# 筛选时间范围内的数据
tp_filtered = tp[time_indices, :, :]
time_dates_filtered = time_dates_utc[time_indices]

# 找到经纬度范围内的索引
lon_indices = np.where((longitude >= lon_min) & (longitude <= lon_max))[0]
lat_indices = np.where((latitude >= lat_min) & (latitude <= lat_max))[0]

# 筛选经纬度范围内的数据
tp_filtered = tp_filtered[:, lat_indices, :][:, :, lon_indices]
filtered_latitude = latitude[lat_indices]
filtered_longitude = longitude[lon_indices]

# 创建一个ExcelWriter对象
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    for t_index, utc_date in enumerate(time_dates_filtered):
        # 转换为 UTC+8 时间
        local_date = utc_date + time_offset

        # 获取当前时间步的数据
        tp_data = tp_filtered[t_index, :, :]

        # 创建DataFrame
        df = pd.DataFrame(tp_data, index=filtered_latitude, columns=filtered_longitude)

        # 创建工作表名称
        sheet_name = local_date.strftime('%Y-%m-%d_%H-%M-%S')  # 格式化为字符串

        # 将DataFrame写入不同的工作表
        df.to_excel(writer, sheet_name=sheet_name)

print(f'数据已成功保存到 {excel_file_path}')

# 创建一个 Pandas DataFrame 用于汇总
df_summary = pd.DataFrame()

# 加载工作簿
wb = load_workbook(input_file, data_only=True)

def format_sheet_name(sheet_name):
    try:
        # 将工作表名称中的下划线替换为冒号
        formatted_name = sheet_name.replace('_', ':')
        # 将格式化的字符串解析为日期时间对象
        dt = datetime.strptime(formatted_name, '%Y-%m-%d:%H-%M-%S')
        # 将日期时间对象格式化为所需的字符串
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except ValueError:
        # 如果转换失败，则返回原始工作表名称
        return sheet_name

# 收集所有工作表的数据
all_data = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    data = {first_column_title: format_sheet_name(sheet_name)}  # 使用自定义标题
    for cell, label in cell_locations.items():
        cell_value = ws[cell].value
        data[label] = cell_value
    all_data.append(data)

# 创建 DataFrame
df_summary = pd.DataFrame(all_data)

# 使用 openpyxl 创建新的工作簿
wb_summary = Workbook()
ws_summary = wb_summary.active
ws_summary.title = "Summary"

# 将 DataFrame 转换为行并添加到工作表
for row in dataframe_to_rows(df_summary, index=False, header=True):
    ws_summary.append(row)

# 设置列宽
for i, column in enumerate(df_summary.columns):
    col_letter = chr(65 + i)  # 获取列字母，例如 A、B、C 等
    if i == 0:
        ws_summary.column_dimensions[col_letter].width = 20
    else:
        ws_summary.column_dimensions[col_letter].width = 12.5

# 保存到 Excel 文件
wb_summary.save(output_file)

print(f"指定单元格的数据已成功汇总并保存到 {output_file}")

